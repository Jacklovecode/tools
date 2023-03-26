# -*- coding: utf-8 -*-
"""
  该脚本用于获取ip对应的域名。
"""

import argparse
import os
import sys
import time
from lxml import etree
import requests
import socket
import re
import tldextract
import openpyxl as xl
from openpyxl.styles import Alignment

requests.packages.urllib3.disable_warnings()

#传入ip可获取当前解析domain
#传入domain可获取历史绑定ip
#可获取ipPostion

headers = {
    "user-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36",
    "Connection": "close"
}

def getIpList(domain): # 获取域名解析出的IP列表
  ip_list = []
  try:
    addrs = socket.getaddrinfo(domain, None)
    for item in addrs:
      if item[4][0] not in ip_list:
        ip_list.append(item[4][0])
  except Exception as e:
    pass
  return ip_list

def getDomain(ip,replayNun=0):
    allData = []#爬取反查域名信息
    domainList = []#最终反查域名信息
    ipPosition = []#获取ip位置信息
    histryIp = []#历史绑定ip
    argIsDoamin = False#参数默认非domain
    try:
        req1 = requests.get(url=f"https://site.ip138.com/{ip}/", headers=headers, timeout=10, verify=False)
        if req1.status_code!=200 and replayNun < 2:
            replayNun += 1
            return getDomain(ip,replayNun)
        if req1.status_code != 200 and replayNun == 2:
            domainList.append(f"NtError c:{req1.status_code}")
            return domainList
        html=etree.HTML(req1.text,etree.HTMLParser())
        if re.match(r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$", ip):
            allData = html.xpath('//ul[@id="list"]/li/a[@target="_blank"]/text()') #获取a节点下的内容,获取到ip曾解析到的domain   存在老旧数据
        else:
            argIsDoamin = True
            histryIp = html.xpath('//div[@id="J_ip_history"]/p/a[@target="_blank"]/text()') #获取a节点下的内容,获取到域名解析到的ip   存在老旧数据
            allData.append(ip)
        for domin in allData:
            #确保反查到的域名可解析到当前ip   剔除老旧数据
            if argIsDoamin or ( ip in getIpList(domin) ):
                #剔除相同域名
                domainObj = tldextract.extract(domin)
                domainData = f"{domainObj.domain}.{domainObj.suffix}"
                if domainData not in domainList:
                    domainList.append(domainData)
        ipPosition=html.xpath('//div[@class="result result2"]/h3/text()')  #获取ip位置信息
    except Exception as e:
#         print(f"\033[31m[Error] url:https://site.ip138.com/{ip}/ {e}\033[0m")
        domainList.append("NtError")
        pass
    return domainList

def getContent(filePath):
    tmpList = []
    with open(filePath,"r") as file:
        for ip in file.readlines():
            tmpList.append(ip)
    return tmpList


def fileQueryDomain(filePath):
    tmpList = getContent(filePath)
    ip_domain = []
    for ip in tmpList:
        ip = ip.replace('\n', '')
        domainList = getDomain(ip)
        if len(domainList) !=0:
            result = ip, domainList[0]
        else:
            result = ip, ''
        ip_domain.append(result)
        print(ip_domain)
        time.sleep(2)
    data_saver(ip_domain)

def singleIp(ip):
    ip_domain = []
    ip = ip.replace('\n', '')
    domainList = getDomain(ip)
    if len(domainList) != 0:
        result = ip, domainList[0]
    else:
        result = ip, ''
    ip_domain.append(result)
    print("{}:{}".format(ip, domainList))
    data_saver(ip_domain)

def data_saver(ip_doamin):
    """
    打印最终结果，并保存数据至Excel表格，同时调整表格格式。
    """
    total_row = len(ip_doamin)
    if total_row == 1:
        total_row = 0
    elif total_row == 0:
        return print("所查ip无对应domain\n")
    # Windows获取桌面路径，将表格保存到桌面，其他系统默认保存到/home/文件夹下
    if os.name == "nt":
        import winreg
        # 用户更改过桌面路径，则需获取User Shell Folders才能获取到准确的桌面路径，否则不会保存到实际的桌面
        subkey = r'Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders'
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, subkey, 0)
        desktop_raw = str(winreg.QueryValueEx(key, "Desktop")[0])
        if desktop_raw == "%USERPROFILE%\Desktop":
            # 此时情况为用户未更改过桌面路径，则需获取系统默认路径
            subkey = r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, subkey, 0)
            desktop_raw = str(winreg.QueryValueEx(key, "Desktop")[0])
        desktop_path = desktop_raw.replace('\\', '/') + "/"
        file_path = f"{desktop_path}ip对应domain.xlsx"
    else:
        file_path = '/home/ip对应domain.xlsx'
    # 存在对应文件，则读取表格追加写入，不存在则创建，并设置表格的标题、列宽、冻结窗格、文字布局等格式
    if os.path.exists(file_path):
        wb = xl.load_workbook(file_path)
        ws = wb['ip对应domain']
        max_row = ws.max_row
        start = max_row + 1
        total_row = total_row + start
        after_title = 0
    else:
        wb = xl.Workbook()
        ws = wb.active
        ws.title = "ip对应domain"
        title_list = ['ip', '域名']
        for i in range(0, 2):
            ws.cell(1, i + 1).value = title_list[i]
        col_width = {'A': 45, 'B': 40}
        for k, v in col_width.items():
            ws.column_dimensions[k].width = v
        ws.freeze_panes = 'A2'
        start = 0
        after_title = 2
    # 写入查询数据
    for j in range(start, total_row + 1):
        for k in range(0, 2):
            try:
                ws.cell(j + after_title, k + 1).value = ip_doamin[j - start][k]
            except:
                continue
    # 垂直居中
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            ws.cell(row + 1, col + 1).alignment = Alignment(horizontal='center', vertical='center')
    try:
        wb.save(file_path)
    except PermissionError:
        print("** ip对应domain表格已打开，无法写入文件。如需写入，请关闭文件后重新执行！ **\n")
        return -1
    print(f"查询结果保存在：{file_path}\n")
    return 'OK'

def parese_args():
    parse = argparse.ArgumentParser(epilog='\tExample: \r\npython '+sys.argv[0]+" -f example.txt"
                                    , description="查询ip地址对应的域名")
    parse.add_argument("-t", "--target", help="查询单个ip对应的域名")
    parse.add_argument("-f", "--file", help="查询文件中所有ip地址对应的域名")
    return parse.parse_args()

if __name__ == '__main__':
    args = parese_args()
    if args.file !=None:
        fileQueryDomain(args.file)
    elif args.file ==None:
        if args.target!=None:
            singleIp(args.target)
    else:
        sys.exit(0)
# print(getDomain("110.242.68.66"))
# print(getDomain("minio-api.lesso.com"))